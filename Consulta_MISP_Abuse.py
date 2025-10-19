import os
import time
import json
import ipaddress
import urllib3
import requests
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel, Label
from tkinter.ttk import Progressbar
from email.utils import parsedate_to_datetime
from datetime import datetime, timezone
import socket
import concurrent.futures
import re
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# =========================
# Configurações
# =========================
cache_file = 'ips_cache.json'

# TTL (segundos) por serviço – controla o "frescor" do cache
CACHE_TTL = {
    'virustotal':    24 * 3600,  # 24h
    'abuseip':       12 * 3600,  # 12h
    'ipqualityscore':12 * 3600,  # 12h
    'misp':           3 * 3600,  # 3h  (instância XXXXXX)
    'misp_ics':       1 * 3600,  # 1h  
    'rdns':          7  * 24 * 3600,  # 7 dias
}

# Ignorar cache nesta execução (force refresh)
FORCE_REFRESH = os.getenv('FORCE_REFRESH', '0') == '1'

# Tempo máximo para aguardar Retry-After e re-tentar imediatamente
MAX_BLOCK_WAIT_SECONDS = 15

# Timeout do reverse DNS (segundos)
RDNS_TIMEOUT = 3

# Estatísticas por serviço (métricas e confirmação de consultas)
SERVICE_STATS = {
    'virustotal':   {'attempts': 0, 'ok': 0, 'rate_limited': 0, 'errors': 0, 'skipped': 0, 'reactivations': 0},
    'abuseip':      {'attempts': 0, 'ok': 0, 'rate_limited': 0, 'errors': 0, 'skipped': 0, 'reactivations': 0},
    'ipqualityscore':{'attempts': 0,'ok': 0, 'rate_limited': 0, 'errors': 0, 'skipped': 0, 'reactivations': 0},
    'misp':         {'attempts': 0, 'ok': 0, 'rate_limited': 0, 'errors': 0, 'skipped': 0, 'reactivations': 0},
    'misp_ics':     {'attempts': 0, 'ok': 0, 'rate_limited': 0, 'errors': 0, 'skipped': 0, 'reactivations': 0},
}

# Próxima janela permitida (pós 429) por serviço
SERVICE_NEXT_ALLOWED = {k: 0.0 for k in SERVICE_STATS.keys()}

# =========================
# Sessão HTTP (pool + retries transitórios)
# =========================
SESSION = requests.Session()
SESSION.headers.update({
    'User-Agent': 'SOC-BH-IP-Checker/1.1',
    'Accept': 'application/json'
})
# Retries apenas para falhas transitórias 5xx (NÃO inclui 429, que tratamos manualmente)
retry_strategy = Retry(
    total=3,
    status_forcelist=[500, 502, 503, 504],
    allowed_methods=["GET", "POST"],
    backoff_factor=0.5,
    raise_on_status=False,
)
adapter = HTTPAdapter(max_retries=retry_strategy, pool_connections=20, pool_maxsize=50)
SESSION.mount("https://", adapter)
SESSION.mount("http://", adapter)

# =========================
# Utilidades
# =========================
def _now() -> int:
    return int(time.time())

def escolher_arquivo():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title='Selecione o arquivo Excel com os IPs',
        filetypes=[('Excel files', '*.xlsx')]
    )

def is_public_ip(ip):
    try:
        return ipaddress.ip_address(ip).is_global
    except ValueError:
        return False

def carregar_cache():
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError:
            return {}
    return {}

def salvar_cache(cache):
    with open(cache_file, 'w', encoding='utf-8') as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)

def mostrar_progresso(total_ips):
    progress_window = Toplevel()
    progress_window.title("Progresso da Consulta")
    Label(progress_window, text="Consultando IPs...").pack(pady=10)
    progress = Progressbar(progress_window, orient='horizontal', length=300, mode='determinate')
    progress.pack(pady=20)
    return progress_window, progress

def atualizar_progresso(progress, current, total):
    progress['value'] = (current / total) * 100
    progress.update()

def criar_blacklist(ips_blacklist, folder_path):
    if ips_blacklist:
        out = os.path.join(folder_path, 'Blacklist.xlsx')
        pd.DataFrame({'IP': ips_blacklist}).to_excel(out, sheet_name='Blacklist', index=False)
        print(f"IPs da Blacklist salvos em {out}")

def parse_retry_after(value):
    """Aceita segundos (string/int) ou data HTTP (ex.: 'Wed, 21 Oct 2015 07:28:00 GMT')."""
    if not value:
        return None
    try:
        return int(value)
    except Exception:
        pass
    try:
        dt = parsedate_to_datetime(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        now = datetime.now(timezone.utc)
        secs = (dt - now).total_seconds()
        return int(secs) if secs > 0 else 0
    except Exception:
        return None

def schedule_next_allowed(service_name, retry_after):
    secs = parse_retry_after(retry_after)
    wait_secs = secs if secs is not None else 3600  # conservador se não informaram Retry-After
    SERVICE_NEXT_ALLOWED[service_name] = max(SERVICE_NEXT_ALLOWED[service_name], time.time() + wait_secs)
    return wait_secs

def maybe_wait_until_allowed(service_name):
    """Se ainda não pode chamar, espera até MAX_BLOCK_WAIT_SECONDS; caso contrário marca 'skipped'."""
    now = time.time()
    next_allowed = SERVICE_NEXT_ALLOWED.get(service_name, 0.0)
    if now >= next_allowed:
        return True
    wait = next_allowed - now
    if wait <= MAX_BLOCK_WAIT_SECONDS:
        print(f"[{service_name}] Aguardando {int(wait)}s (Retry-After) para reativar automaticamente...")
        time.sleep(wait)
        SERVICE_STATS[service_name]['reactivations'] += 1
        return True
    else:
        SERVICE_STATS[service_name]['skipped'] += 1
        return False

def call_with_rate_limit(service_name, call_fn):
    """
    Envolve a chamada ao serviço com gestão de rate limit:
    - Respeita SERVICE_NEXT_ALLOWED (pós 429).
    - Em caso de 429, agenda próxima janela e tenta 1 retry se o tempo for curto.
    Retorna: {'status': 'ok'|'rate_limited'|'error'|'skipped', 'value':..., 'retry_after':...}
    """
    if not maybe_wait_until_allowed(service_name):
        return {'status': 'skipped', 'value': 'Pulado (rate limit)'}

    SERVICE_STATS[service_name]['attempts'] += 1
    res = call_fn()

    if res.get('status') == 'rate_limited':
        SERVICE_STATS[service_name]['rate_limited'] += 1
        wait_secs = schedule_next_allowed(service_name, res.get('retry_after'))
        if wait_secs <= MAX_BLOCK_WAIT_SECONDS:
            print(f"[{service_name}] Rate-limited. Esperando {wait_secs}s para tentar novamente...")
            time.sleep(max(0, wait_secs))
            SERVICE_STATS[service_name]['reactivations'] += 1
            SERVICE_STATS[service_name]['attempts'] += 1
            res = call_fn()
            if res.get('status') == 'rate_limited':
                SERVICE_STATS[service_name]['rate_limited'] += 1
        return res
    elif res.get('status') == 'ok':
        SERVICE_STATS[service_name]['ok'] += 1
        return res
    else:
        SERVICE_STATS[service_name]['errors'] += 1
        return res

# =========================
# Cache helpers com TTL
# =========================
def cache_get_fresh(cache: dict, ip: str, key: str, ts_key: str, ttl_seconds: int):
    item = cache.get(ip)
    if not item:
        return None
    if FORCE_REFRESH:
        return None
    val = item.get(key)
    ts  = item.get(ts_key)
    if val is None or ts is None:
        return None
    if (_now() - ts) <= ttl_seconds:
        return val
    return None

def cache_put_ok(cache: dict, ip: str, key: str, ts_key: str, value):
    """Grava no cache com timestamp apenas para resultados válidos (status 'ok')."""
    item = cache.get(ip) or {}
    item[key] = value
    item[ts_key] = _now()
    cache[ip] = item

# =========================
# Reverse DNS (PTR)
# =========================
def reverse_dns_lookup(ip, timeout=RDNS_TIMEOUT):
    """Retorna o FQDN do PTR; None se não houver/erro/timeout."""
    def _lookup():
        try:
            host, alias, addrs = socket.gethostbyaddr(ip)
            return host.strip('.')
        except Exception:
            return None

    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(_lookup)
        try:
            return fut.result(timeout=timeout)
        except concurrent.futures.TimeoutError:
            return None

# =========================
# Integrações (APIs) usando SESSION
# =========================
def verificar_ip_virustotal(ip, api_key, timeout=10):
    url = f'https://www.virustotal.com/api/v3/ip_addresses/{ip}'
    headers = {'x-apikey': api_key}
    try:
        resp = SESSION.get(url, headers=headers, timeout=timeout)
        if resp.status_code == 429:
            return {'status': 'rate_limited', 'retry_after': resp.headers.get('Retry-After')}
        resp.raise_for_status()
        data = resp.json()
        malicious = data.get('data', {}).get('attributes', {}).get('last_analysis_stats', {}).get('malicious', 0)
        return {'status': 'ok', 'value': malicious}
    except requests.RequestException as e:
        print(f"[virustotal] Erro IP {ip}: {e}")
        return {'status': 'error', 'error': str(e)}

def verificar_ip_abuseip(ip, api_key, timeout=10):
    url = 'https://api.abuseipdb.com/api/v2/check'
    headers = {'Accept': 'application/json', 'Key': api_key}
    params = {'ipAddress': ip, 'maxAgeInDays': 90}
    try:
        resp = SESSION.get(url, headers=headers, params=params, timeout=timeout)
        if resp.status_code == 429:
            return {'status': 'rate_limited', 'retry_after': resp.headers.get('Retry-After')}
        resp.raise_for_status()
        data = resp.json()
        score = data.get('data', {}).get('abuseConfidenceScore', 0)
        return {'status': 'ok', 'value': score}
    except requests.RequestException as e:
        print(f"[abuseip] Erro IP {ip}: {e}")
        return {'status': 'error', 'error': str(e)}

def verificar_ip_ipqualityscore(ip, api_key, timeout=10):
    url = f'https://www.ipqualityscore.com/api/json/ip/{api_key}/{ip}'
    try:
        resp = SESSION.get(url, timeout=timeout)
        if resp.status_code == 429:
            return {'status': 'rate_limited', 'retry_after': resp.headers.get('Retry-After')}
        resp.raise_for_status()
        data = resp.json()
        fraud_score = data.get('fraud_score', 0)
        return {'status': 'ok', 'value': fraud_score}
    except requests.RequestException as e:
        print(f"[ipqualityscore] Erro IP {ip}: {e}")
        return {'status': 'error', 'error': str(e)}

# MISP legado (XXXXXXX) – verify=False conforme original
def verificar_ip_misp(ip, timeout=10):
    url = "https://XXXXXXX/attributes/restSearch"
    headers = {
        "Authorization": "ki98bT7IkQgrua5Zn4VkDvUcm8GP8bxEmJcgqlCl",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    data = {"type": "ip-src", "value": ip, "to_ids": True}
    try:
        resp = SESSION.post(url, headers=headers, json=data, verify=False, timeout=timeout)
        if resp.status_code == 429:
            return {'status': 'rate_limited', 'retry_after': resp.headers.get('Retry-After')}
        if resp.status_code == 200:
            ip_data = resp.json()
            is_malicious = bool(ip_data.get('response', {}).get('Attribute'))
            return {'status': 'ok', 'value': is_malicious}
        return {'status': 'error', 'error': f'Status {resp.status_code}'}
    except requests.RequestException as e:
        print(f"[misp] Erro IP {ip}: {e}")
        return {'status': 'error', 'error': str(e)}

# MISP ICS-CSIRT
def verificar_usuario_misp_ics(api_key, timeout=10):
    url = "https://misp.ics-csirt.io/users/view/me"
    headers = {"Authorization": api_key, "Content-Type": "application/json"}
    try:
        resp = SESSION.get(url, headers=headers, timeout=timeout)
        if resp.status_code == 429:
            return {'status': 'rate_limited', 'retry_after': resp.headers.get('Retry-After')}
        resp.raise_for_status()
        data = resp.json()
        user_info = {}
        try:
            user = data.get('response', {}).get('User', {})
            user_info = {
                'id': user.get('id'),
                'email': user.get('email'),
                'org_id': user.get('org_id'),
                'role_id': user.get('role_id'),
                'disabled': user.get('disabled'),
            }
        except Exception:
            pass
        return {'status': 'ok', 'value': user_info if user_info else data}
    except requests.HTTPError:
        try:
            err_body = resp.json()
        except Exception:
            err_body = resp.text
        return {'status': 'error', 'error': f'HTTP {resp.status_code}: {err_body}'}
    except requests.RequestException as e:
        return {'status': 'error', 'error': str(e)}

def verificar_ip_misp_ics(ip, api_key, timeout=10, verify_tls=True):
    url = "https://misp.ics-csirt.io/attributes/restSearch"
    headers = {"Authorization": api_key, "Content-Type": "application/json"}
    payload = {
        "type": ["ip-src", "ip-dst"],
        "value": ip,
        "to_ids": True,
        "returnFormat": "json"
    }
    try:
        resp = SESSION.post(url, headers=headers, json=payload, timeout=timeout, verify=verify_tls)
        if resp.status_code == 429:
            return {'status': 'rate_limited', 'retry_after': resp.headers.get('Retry-After')}
        resp.raise_for_status()
        data = resp.json()
        attrs = data.get('response', {}).get('Attribute') or []
        if isinstance(attrs, dict):
            attrs = [attrs]
        is_malicious = bool(attrs)
        return {'status': 'ok', 'value': is_malicious}
    except requests.HTTPError:
        try:
            err_body = resp.json()
        except Exception:
            err_body = resp.text
        return {'status': 'error', 'error': f'HTTP {resp.status_code}: {err_body}'}
    except requests.RequestException as e:
        return {'status': 'error', 'error': str(e)}

# =========================
# Helpers para planilha específica
# =========================
TARGET_SHEETS = [
    'Talking with External IPs',
    'Talking with Ghost Assets'
]

NORMALIZED_TARGETS = [re.sub('[^A-Za-z]+', '', s).lower() for s in TARGET_SHEETS]

RESULT_COLUMNS = [
    'Dominio_RDNS', 'VirusTotal_Score', 'AbuseIPDB_Reported', 'IPQualityScore_Fraud_Score',
    'MISP_Malicious', 'MISP_ICS_Malicious', 'VirusTotal_Malicious', 'AbuseIPDB_Malicious', 'IPQualityScore_Malicious'
]


def normalize_sheet_name(name):
    return re.sub('[^A-Za-z]+', '', str(name)).lower()


def find_header_cell(ws, header_value):
    """Procura a célula (row, col) cujo valor textual case-insensitive corresponde a header_value."""
    hv = str(header_value).strip().lower()
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            try:
                if cell.value is None:
                    continue
                if str(cell.value).strip().lower() == hv:
                    return cell.row, cell.column
            except Exception:
                continue
    return None, None


def collect_ips_below(ws, start_row, col):
    ips = []
    r = start_row + 1
    while True:
        cell = ws.cell(row=r, column=col)
        val = cell.value
        if val is None or (isinstance(val, str) and val.strip() == ''):
            break
        ips.append((str(val).strip(), r))
        r += 1
    return ips


def find_first_blank_col_to_right(ws, row, start_col):
    c = start_col + 1
    # procura até 200 colunas à direita para cautela
    for offset in range(0, 200):
        col = c + offset
        cell = ws.cell(row=row, column=col)
        if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ''):
            return col
    return None

# =========================
# Principal
# =========================

def main():
    file_path = escolher_arquivo()
    if not file_path:
        print("Nenhum arquivo selecionado.")
        return

    folder_path = os.path.dirname(file_path)

    # Chaves de API (mantidas) + ICS-CSIRT
    api_keys = {
        'virustotal': '',
        'abuseip': '',
        'ipqualityscore': '',
        'misp_ics': ''
    }

    # Checagem de chaves
    for service, key in api_keys.items():
        if key.startswith('SUA_CHAVE'):
            messagebox.showwarning("Chave API Ausente", f"Por favor, configure a chave API para {service}.")
            return

    # Autenticação no MISP ICS-CSIRT (sem cache)
    print("Verificando acesso ao MISP ICS-CSIRT...")
    me_res = call_with_rate_limit('misp_ics', lambda: verificar_usuario_misp_ics(api_keys['misp_ics']))
    if me_res['status'] == 'ok':
        ui = me_res.get('value', {})
        user_str = (f"id={ui.get('id')}, email={ui.get('email')}, org_id={ui.get('org_id')}, "
                    f"role_id={ui.get('role_id')}, disabled={ui.get('disabled')}")
        print(f"[misp_ics] Autenticado: {user_str}")
        try:
            messagebox.showinfo("MISP ICS-CSIRT", f"Autenticado no MISP ICS-CSIRT.\n{user_str}")
        except Exception:
            pass
    elif me_res['status'] == 'rate_limited':
        ra = me_res.get('retry_after') or 'desconhecido'
        print(f"[misp_ics] Rate-limited no /users/view/me (Retry-After={ra}).")
    else:
        print(f"[misp_ics] Falha ao consultar /users/view/me: {me_res.get('error')}")

    # Carregar cache
    cache = carregar_cache()

    # Abrir workbook com openpyxl para escrita célula-a-célula
    wb = load_workbook(filename=file_path)

    # Coleção de tuples (sheet_name, ip, ip_row, ip_col)
    ips_to_query = []

    # Varre as planilhas procurando as targets (ignorando números e chars)
    for sheet_name in wb.sheetnames:
        norm = normalize_sheet_name(sheet_name)
        if norm in NORMALIZED_TARGETS:
            ws = wb[sheet_name]
            # decidir header a procurar conforme o tipo
            if normalize_sheet_name(sheet_name) == re.sub('[^A-Za-z]+', '', TARGET_SHEETS[0]).lower():
                header_text = 'External IP'
            else:
                header_text = 'Ip'

            header_row, header_col = find_header_cell(ws, header_text)
            if not header_row:
                print(f"Cabeçalho '{header_text}' não encontrado em '{sheet_name}', pulando.")
                continue

            # coletar ips na coluna abaixo
            collected = collect_ips_below(ws, header_row, header_col)
            if not collected:
                print(f"Nenhum IP encontrado abaixo de '{header_text}' em '{sheet_name}'.")
                continue

            # identificar primeira coluna em branco à direita do cabeçalho para colocar títulos de resultado
            result_start_col = find_first_blank_col_to_right(ws, header_row, header_col)
            if not result_start_col:
                print(f"Não foi possível encontrar coluna em branco para escrever resultados em '{sheet_name}'.")
                continue

            # escreve títulos dos resultados a partir dessa coluna (apenas se vazio)
            for i, col_name in enumerate(RESULT_COLUMNS):
                cell = ws.cell(row=header_row, column=result_start_col + i)
                if cell.value is None or str(cell.value).strip() == '':
                    cell.value = col_name

            # adicionar IPs encontrados à lista geral para consulta
            for ip_val, r in collected:
                ips_to_query.append((sheet_name, ip_val, r, header_row, header_col, result_start_col))

    if not ips_to_query:
        messagebox.showinfo("Nenhum IP", "Não foram encontrados IPs para consultar nas abas alvo.")
        return

    # UI de progresso
    progress_window, progress = mostrar_progresso(len(ips_to_query))

    ips_blacklist = []
    processed = 0

    # Processa cada IP (mantendo lógica original)
    for entry in ips_to_query:
        sheet_name, ip, ip_row, header_row, header_col, result_start_col = entry
        ws = wb[sheet_name]

        processed += 1

        if not is_public_ip(ip):
            # escreve marcadores para IP privado nas colunas de resultado
            ws.cell(row=ip_row, column=result_start_col).value = 'N/A (IP Privado)'
            ws.cell(row=ip_row, column=result_start_col+1).value = 'IP Privado'
            ws.cell(row=ip_row, column=result_start_col+2).value = 'IP Privado'
            ws.cell(row=ip_row, column=result_start_col+3).value = 'IP Privado'
            ws.cell(row=ip_row, column=result_start_col+4).value = 'Não'
            ws.cell(row=ip_row, column=result_start_col+5).value = 'Não'
            ws.cell(row=ip_row, column=result_start_col+6).value = 'Não'
            ws.cell(row=ip_row, column=result_start_col+7).value = 'Não'
            ws.cell(row=ip_row, column=result_start_col+8).value = 'Não'
            atualizar_progresso(progress, processed, len(ips_to_query))
            continue

        # Objeto existente no cache (pode conter dados variados/antigos)
        resultado_ip = cache.get(ip, {}).copy()

        # --- RDNS com TTL
        rdns_cached = cache_get_fresh(cache, ip, 'Dominio_RDNS', 'rdns_ts', CACHE_TTL['rdns'])
        if rdns_cached is not None:
            rdns_value = rdns_cached
        else:
            rdns = reverse_dns_lookup(ip)
            rdns_value = rdns if rdns else 'Sem PTR'
            cache_put_ok(cache, ip, 'Dominio_RDNS', 'rdns_ts', rdns_value)

        # --- VirusTotal com TTL
        vt_cached = cache_get_fresh(cache, ip, 'VirusTotal_Score', 'vt_ts', CACHE_TTL['virustotal'])
        if vt_cached is not None:
            vt_value = vt_cached
        else:
            vt_res = call_with_rate_limit('virustotal', lambda: verificar_ip_virustotal(ip, api_keys['virustotal']))
            if vt_res['status'] == 'ok':
                vt_value = vt_res['value']
                cache_put_ok(cache, ip, 'VirusTotal_Score', 'vt_ts', vt_value)
            elif vt_res['status'] in ('rate_limited', 'skipped'):
                vt_value = 'Pulado (rate limit)' if vt_res['status'] == 'skipped' else 'Rate Limit'
            else:
                vt_value = None

        # --- AbuseIPDB com TTL
        abuse_cached = cache_get_fresh(cache, ip, 'AbuseIPDB_Reported', 'abuse_ts', CACHE_TTL['abuseip'])
        if abuse_cached is not None:
            abuse_value = abuse_cached
        else:
            abuse_res = call_with_rate_limit('abuseip', lambda: verificar_ip_abuseip(ip, api_keys['abuseip']))
            if abuse_res['status'] == 'ok':
                abuse_value = abuse_res['value']
                cache_put_ok(cache, ip, 'AbuseIPDB_Reported', 'abuse_ts', abuse_value)
            elif abuse_res['status'] in ('rate_limited', 'skipped'):
                abuse_value = 'Pulado (rate limit)' if abuse_res['status'] == 'skipped' else 'Rate Limit'
            else:
                abuse_value = None

        # --- IPQualityScore com TTL
        ipq_cached = cache_get_fresh(cache, ip, 'IPQualityScore_Fraud_Score', 'ipq_ts', CACHE_TTL['ipqualityscore'])
        if ipq_cached is not None:
            ipq_value = ipq_cached
        else:
            ipq_res = call_with_rate_limit('ipqualityscore', lambda: verificar_ip_ipqualityscore(ip, api_keys['ipqualityscore']))
            if ipq_res['status'] == 'ok':
                ipq_value = ipq_res['value']
                cache_put_ok(cache, ip, 'IPQualityScore_Fraud_Score', 'ipq_ts', ipq_value)
            elif ipq_res['status'] in ('rate_limited', 'skipped'):
                ipq_value = 'Pulado (rate limit)' if ipq_res['status'] == 'skipped' else 'Rate Limit'
            else:
                ipq_value = None

        # --- MISP legado com TTL
        misp_cached = cache_get_fresh(cache, ip, 'MISP_Malicious', 'misp_ts', CACHE_TTL['misp'])
        if misp_cached is not None:
            misp_value = misp_cached
        else:
            misp_res = call_with_rate_limit('misp', lambda: verificar_ip_misp(ip))
            if misp_res['status'] == 'ok':
                misp_value = misp_res['value']
                cache_put_ok(cache, ip, 'MISP_Malicious', 'misp_ts', misp_value)
            elif misp_res['status'] in ('rate_limited', 'skipped'):
                misp_value = 'Pulado (rate limit)' if misp_res['status'] == 'skipped' else 'Rate Limit'
            else:
                misp_value = None

        # --- MISP ICS-CSIRT com TTL (CONFIRMADO: consulta por IP)
        misp_ics_cached = cache_get_fresh(cache, ip, 'MISP_ICS_Malicious', 'misp_ics_ts', CACHE_TTL['misp_ics'])
        if misp_ics_cached is not None:
            misp_ics_value = misp_ics_cached
        else:
            print(f"[misp_ics] Consultando atributos para {ip} ...")
            misp_ics_res = call_with_rate_limit('misp_ics', lambda: verificar_ip_misp_ics(ip, api_keys['misp_ics']))
            if misp_ics_res['status'] == 'ok':
                misp_ics_value = misp_ics_res['value']
                cache_put_ok(cache, ip, 'MISP_ICS_Malicious', 'misp_ics_ts', misp_ics_value)
            elif misp_ics_res['status'] in ('rate_limited', 'skipped'):
                misp_ics_value = 'Pulado (rate limit)' if misp_ics_res['status'] == 'skipped' else 'Rate Limit'
            else:
                misp_ics_value = None

        # Consolidar no registro
        resultado_ip = {
            'Dominio_RDNS': rdns_value,
            'VirusTotal_Score': vt_value,
            'AbuseIPDB_Reported': abuse_value,
            'IPQualityScore_Fraud_Score': ipq_value,
            'MISP_Malicious': misp_value,
            'MISP_ICS_Malicious': misp_ics_value
        }
        cache[ip] = cache.get(ip, {})
        cache[ip].update(resultado_ip)

        # Interpretação
        vt_score = resultado_ip.get('VirusTotal_Score')
        abuse_score = resultado_ip.get('AbuseIPDB_Reported')
        ipq_score = resultado_ip.get('IPQualityScore_Fraud_Score')
        misp_flag = resultado_ip.get('MISP_Malicious')
        misp_ics_flag = resultado_ip.get('MISP_ICS_Malicious')

        vt_mal = 'Sim' if isinstance(vt_score, (int, float)) and vt_score > 3 else 'Não'
        abuse_mal = 'Sim' if isinstance(abuse_score, (int, float)) and abuse_score > 0 else 'Não'
        ipq_mal = 'Sim' if isinstance(ipq_score, (int, float)) and ipq_score > 0.5 else 'Não'
        misp_mal = 'Sim' if isinstance(misp_flag, bool) and misp_flag else 'Não'
        misp_ics_mal = 'Sim' if isinstance(misp_ics_flag, bool) and misp_ics_flag else 'Não'

        # Escrever resultados nas colunas apropriadas
        ws.cell(row=ip_row, column=result_start_col).value = resultado_ip.get('Dominio_RDNS', 'Sem PTR')
        ws.cell(row=ip_row, column=result_start_col+1).value = vt_score
        ws.cell(row=ip_row, column=result_start_col+2).value = abuse_score
        ws.cell(row=ip_row, column=result_start_col+3).value = ipq_score
        ws.cell(row=ip_row, column=result_start_col+4).value = 'Sim' if misp_mal == 'Sim' else (misp_flag if isinstance(misp_flag, str) else 'Não')
        ws.cell(row=ip_row, column=result_start_col+5).value = 'Sim' if misp_ics_mal == 'Sim' else (misp_ics_flag if isinstance(misp_ics_flag, str) else 'Não')
        ws.cell(row=ip_row, column=result_start_col+6).value = vt_mal if vt_score not in ('Rate Limit', 'Pulado (rate limit)') else 'Não'
        ws.cell(row=ip_row, column=result_start_col+7).value = abuse_mal if abuse_score not in ('Rate Limit', 'Pulado (rate limit)') else 'Não'
        ws.cell(row=ip_row, column=result_start_col+8).value = ipq_mal if ipq_score not in ('Rate Limit', 'Pulado (rate limit)') else 'Não'

        # Blacklist
        try:
            if (isinstance(vt_score, (int, float)) and vt_score > 3) or \
               (isinstance(abuse_score, (int, float)) and abuse_score > 0) or \
               (isinstance(ipq_score, (int, float)) and ipq_score > 0.5) or \
               (isinstance(misp_flag, bool) and misp_flag) or \
               (isinstance(misp_ics_flag, bool) and misp_ics_flag):
                ips_blacklist.append(ip)
        except Exception:
            pass

        atualizar_progresso(progress, processed, len(ips_to_query))

    # Persistências finais
    salvar_cache(cache)
    progress_window.destroy()

    # Salva workbook sobrescrevendo o original (resultados inseridos nas mesmas abas)
    wb.save(file_path)
    print(f"Resultados gravados de volta em: {file_path}")

    criar_blacklist(ips_blacklist, folder_path)

    # Resumo/Confirmações
    summary_lines = ["\n=== Resumo por serviço ==="]
    for svc, c in SERVICE_STATS.items():
        summary_lines.append(
            f"- {svc}: tentativas={c['attempts']}, ok={c['ok']}, 429={c['rate_limited']}, "
            f"skipped={c['skipped']}, erros={c['errors']}, reativações={c['reactivations']}"
        )
    print("\n".join(summary_lines))

    if SERVICE_STATS['misp_ics']['attempts'] > 0:
        print("[misp_ics] Confirmação: consultas foram realizadas ao misp.ics-csirt.io.")
    else:
        print("[misp_ics] Atenção: nenhuma tentativa de consulta foi registrada.")

    try:
        messagebox.showinfo("Concluído",
            "Consulta de IPs finalizada.\n"
            "Use FORCE_REFRESH=1 para ignorar cache.\n"
            "Resultados foram escritos nas abas originais.")
    except Exception:
        pass

# Execução
if __name__ == '__main__':
    main()
